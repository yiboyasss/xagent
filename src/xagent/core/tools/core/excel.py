from __future__ import annotations

import asyncio
from types import ModuleType
from typing import Any, Dict, List, Optional, Set

# Optional import for openpyxl
openpyxl: ModuleType | None = None
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.comments import Comment
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    openpyxl_not_installed_exception = RuntimeError(
        "openpyxl is not installed. "
        "Install with: pip install 'xagent[document-processing]'"
    )


def calculate_cell_size_from_font(
    font_size: int,
    text_length: Optional[int] = None,
    text: Optional[str] = None,
    max_width: Optional[float] = None,
    padding: float = 2.0,
    wrap_text: bool = True,
) -> dict:
    """
    Automatically calculate recommended cell dimensions based on font size, supporting auto-wrap.
    Args:
        font_size (int): Font size
        text_length (int, optional): Text length for column width calculation
        text (str, optional): Actual text content for wrap calculation
        max_width (float, optional): Maximum column width limit
        padding (float): Padding, defaults to 2.0
        wrap_text (bool): Whether to enable auto-wrap, defaults to True
    Returns:
        dict: Dictionary containing recommended row height and column width
    """
    base_row_height = font_size * 1.2 + padding

    char_width = font_size * 0.4

    if text and wrap_text:
        if max_width:
            column_width = max_width
        else:
            chars_per_line = 15
            column_width = chars_per_line * char_width + padding
    elif text_length:
        column_width = text_length * char_width + padding
    else:
        column_width = font_size * 1.5 + padding
    if text and wrap_text:
        estimated_chars_per_line = (
            int(column_width / char_width * 1.8) if column_width > 0 else 15
        )
        if estimated_chars_per_line > 0:
            estimated_lines = max(
                1,
                (len(text) + estimated_chars_per_line - 1) // estimated_chars_per_line,
            )
            row_height = base_row_height * estimated_lines + padding
        else:
            estimated_lines = 1
            row_height = base_row_height
    else:
        row_height = base_row_height

    return {
        "row_height": round(row_height, 2),
        "column_width": round(column_width, 2),
    }


def _apply_basic_font(cell: Cell, font_size: Optional[int]) -> None:
    if font_size is None:
        return
    cell.font = cell.font.copy(size=font_size)


def _apply_wrap_alignment(cell: Cell, wrap_text: bool) -> None:
    if wrap_text:
        cell.alignment = cell.alignment.copy(wrap_text=True)


def _apply_link_style(cell: Cell, font_size: Optional[int], wrap_text: bool) -> None:
    size_to_use = font_size or 11
    cell.font = Font(color="0563C1", underline="single", size=size_to_use)
    _apply_wrap_alignment(cell, wrap_text)


def _auto_size_if_needed(
    sheet: Worksheet,
    cell: Cell,
    value: Any,
    font_size: Optional[int],
    auto_size: bool,
    max_width: Optional[float],
    padding: float,
    wrap_text: bool,
    updated_rows: Set[int],
    updated_columns: Set[str],
) -> None:
    if not auto_size or font_size is None:
        return
    text_length = len(str(value)) if value is not None else 0
    text_content = str(value) if value is not None else ""
    size_info = calculate_cell_size_from_font(
        font_size,
        text_length,
        text_content,
        max_width,
        padding,
        wrap_text,
    )
    row_number = cell.row
    if row_number not in updated_rows:
        sheet.row_dimensions[row_number].height = size_info["row_height"]
        updated_rows.add(row_number)
    column_letter = get_column_letter(cell.column)
    if column_letter not in updated_columns:
        sheet.column_dimensions[column_letter].width = size_info["column_width"]
        updated_columns.add(column_letter)


async def read_excel_cells(file_path: str, sheet_name: str = "Sheet1") -> List[str]:
    """Read all non-empty cells' positions and content from an Excel file.
    Args:
        file_path (str): Excel file path
        sheet_name (str): Worksheet name, defaults to "Sheet1"
    Returns:
        List[str]: List containing all cell positions and content
    """
    if not openpyxl:
        raise openpyxl_not_installed_exception

    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name]

        cells_info = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    cell_address = cell.coordinate
                    cell_value = str(cell.value).strip()
                    cells_info.append(f"Worksheet {cell_address}: {cell_value}")

        if not cells_info:
            return [f"Worksheet '{sheet_name}' No non-empty cells were found."]
        return cells_info
    except FileNotFoundError:
        raise ValueError(f"File not found: {file_path}")
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}") from e


async def update_excel_cells(
    file_path: str,
    updates: List[Dict[str, Any]],
    sheet_name: str = "Sheet1",
    font_size: Optional[int] = 11,
    auto_size: bool = True,
    padding: float = 2.0,
    wrap_text: bool = True,
    max_width: Optional[float] = None,
    auto_size_links: bool = False,
) -> str:
    """
    Batch update Excel cells, supports hyperlinks and comments By default, do not use hyperlinks and annotations.
    Args:
        file_path (str): Excel file path.
        updates (list): A list of dictionaries, each can contain:
                       - 'cell_address' (str, required): Cell address
                       - 'new_value' (any, required): New value
                       - 'comment_text' (str, optional): Comment content
                       - 'comment_author' (str, optional): Comment author (defaults to 'test')
                       - 'hyperlink' (str, optional): Hyperlink
                       Example: [
                           {"cell_address": "A1", "new_value": "Regular data"},
                           {"cell_address": "A1", "new_value": "Regular data", "comment_text": "This is a comment"},
                           {"cell_address": "B2", "new_value": 12345, "comment_text": "Data needs review", "comment_author": "John"}
                           {"cell_address": "A1", "new_value": "Regular data", "hyperlink": "https://www.google.com"},
                       ]
        sheet_name (str, optional): Worksheet name.
        auto_size (bool, optional): Whether to auto-adjust size.
        padding (float, optional): Padding.
        wrap_text (bool, optional): Whether to auto-wrap.
        max_width (float, optional): Maximum column width.
    Returns:
        str: Result information of the batch update operation.
    """
    if not openpyxl:
        raise openpyxl_not_installed_exception

    loop = asyncio.get_running_loop()

    try:
        workbook = await loop.run_in_executor(None, load_workbook, file_path)
        sheet = workbook[sheet_name]

        updated_count = 0
        updated_rows: set = set()
        updated_columns: set = set()

        for update_item in updates:
            cell_address = update_item.get("cell_address")
            if not cell_address:
                continue

            try:
                cell = sheet[cell_address]

                # Value update
                if "new_value" in update_item:
                    cell.value = update_item.get("new_value")
                new_value = cell.value

                # Hyperlink handling: set or clear if provided
                hyperlink_provided = "hyperlink" in update_item
                hyperlink_target = (
                    update_item.get("hyperlink") if hyperlink_provided else None
                )
                if hyperlink_provided:
                    if hyperlink_target:
                        cell.hyperlink = hyperlink_target
                        _apply_link_style(cell, font_size, wrap_text)
                    else:
                        cell.hyperlink = None
                        _apply_basic_font(cell, font_size)
                        _apply_wrap_alignment(cell, wrap_text)
                else:
                    _apply_basic_font(cell, font_size)
                    _apply_wrap_alignment(cell, wrap_text)

                # Comment handling: set or clear if provided
                if "comment_text" in update_item:
                    comment_text = update_item.get("comment_text")
                    if comment_text:
                        author = update_item.get("comment_author", "Gemini")
                        cell.comment = Comment(comment_text, author)
                    else:
                        cell.comment = None

                # Auto size
                should_auto_size = auto_size and (
                    auto_size_links or not hyperlink_target
                )
                _auto_size_if_needed(
                    sheet,
                    cell,
                    new_value,
                    font_size,
                    should_auto_size,
                    max_width,
                    padding,
                    wrap_text,
                    updated_rows,
                    updated_columns,
                )

                updated_count += 1
            except Exception as e:
                raise ValueError(f"Failed to update cell {cell_address}: {e}") from e

        await loop.run_in_executor(None, workbook.save, file_path)
        return (
            f"Successfully batch updated {updated_count} cells in sheet '{sheet_name}'."
        )
    except Exception as e:
        raise ValueError(f"Batch update Excel failed: {e}") from e
